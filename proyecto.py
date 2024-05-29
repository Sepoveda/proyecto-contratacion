from openpyxl import load_workbook, Workbook
import os
import platform

# Cargar el archivo Excel base
wb_base = load_workbook("CONTRATACION 2024.xlsm")

# Especificar la hoja por su nombre
nombre_hoja = "FESTIVOS"  # Reemplaza "Hoja1" por el nombre de la hoja que necesites
ws_base = wb_base[nombre_hoja]

# Crear un nuevo libro de trabajo y una hoja
wb_nuevo = Workbook()
ws_nuevo = wb_nuevo.active

# Definir el rango de celdas a copiar (de A2 a C5)
rango_celdas = ws_base['A1:C35']

# Iterar sobre las filas del rango y copiar los datos al nuevo archivo
for fila in rango_celdas:
    datos_fila = [celda.value for celda in fila]
    ws_nuevo.append(datos_fila)

# Guardar el nuevo libro
archivo_nuevo = "nuevo.xlsx"
wb_nuevo.save(archivo_nuevo)

# Abrir el archivo automáticamente después de guardarlo
def abrir_archivo(archivo):
    sistema = platform.system()
    if sistema == "Windows":
        os.startfile(archivo)
    elif sistema == "Darwin":  # macOS
        os.system(f"open {archivo}")
    elif sistema == "Linux":
        os.system(f"xdg-open {archivo}")

# Llamar a la función para abrir el archivo
abrir_archivo(archivo_nuevo)
